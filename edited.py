"""
Riccati BVP Explorer — Modern Streamlit GUI
✅ Fully controllable parameters
✅ LaTeX equation rendering
✅ Light/Dark mode toggle
✅ Big graphs in all tabs
"""

import streamlit as st
import numpy as np
import matplotlib.pyplot as plt
import math
import os
from datetime import datetime
import io

# Excel export dependency
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from riccati_core import (
    solve_combined, exact_combined, solve_branch, 
    y_exact, yprime_exact, yprime_from_y
)

# Page configuration
st.set_page_config(
    page_title="Riccati BVP Explorer",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state for theme
if 'theme' not in st.session_state:
    st.session_state.theme = 'dark'

# Theme toggle in sidebar
with st.sidebar:
    st.header("⚙️ Settings")
    theme_mode = st.selectbox("Theme", ["Dark", "Light"], 
                              index=0 if st.session_state.theme == 'dark' else 1,
                              key="theme_selector")
    st.session_state.theme = theme_mode.lower()

# Theme-dependent colors and styles
if st.session_state.theme == 'dark':
    bg_color = '#2b2b2b'
    axes_bg = '#2b2b2b'
    text_color = '#e8e8e8'
    grid_color = '#444444'
    accent_color = '#00d4ff'
    header_color = '#00d4ff'
    card_bg = '#3a3a3a'
    border_color = '#555555'
else:
    bg_color = '#ffffff'
    axes_bg = '#ffffff'
    text_color = '#1F2937'
    grid_color = '#CCCCCC'
    accent_color = '#0066CC'
    header_color = '#0066CC'
    card_bg = '#F5F5F5'
    border_color = '#DDDDDD'

# Custom CSS based on theme
st.markdown(f"""
    <style>
    .main-header {{
        font-size: 1.8rem;
        font-weight: bold;
        color: {header_color};
        text-align: center;
        margin-bottom: 0.3rem;
    }}
    .sub-header {{
        font-size: 0.95rem;
        color: #888888;
        text-align: center;
        margin-bottom: 1rem;
    }}
    .equation-box {{
        background: {'linear-gradient(135deg, #1e1e1e 0%, #2b2b2b 100%)' if st.session_state.theme == 'dark' else 'linear-gradient(135deg, #F5F5F5 0%, #FFFFFF 100%)'};
        padding: 0.8rem;
        border-radius: 8px;
        border: 1px solid {border_color};
        margin: 0.5rem 0;
        text-align: center;
    }}
    .info-box {{
        background-color: {card_bg};
        padding: 1rem;
        border-radius: 8px;
        border: 1px solid {border_color};
        margin: 1rem 0;
        font-family: 'Consolas', monospace;
        font-size: 0.9rem;
        color: {text_color};
    }}
    </style>
""", unsafe_allow_html=True)

# Color palettes - different colors for each graph
COLOR_PALETTES = {
    'euler': ['#2563EB', '#3B82F6', '#60A5FA'],  # Blue shades
    'rk4': ['#F59E0B', '#F97316', '#FB923C'],    # Orange/Amber shades
    'rk45': ['#10B981', '#059669', '#34D399'],   # Green/Emerald shades
    'exact': '#111827' if st.session_state.theme == 'dark' else '#000000',
    'sing': '#DC2626',  # Red for singularity
}

# Configure matplotlib based on theme
if st.session_state.theme == 'dark':
    plt.style.use('dark_background')
else:
    plt.style.use('default')

plt.rcParams.update({
    'figure.facecolor': bg_color,
    'axes.facecolor': axes_bg,
    'axes.edgecolor': border_color,
    'axes.labelcolor': text_color,
    'text.color': text_color,
    'xtick.color': text_color,
    'ytick.color': text_color,
    'grid.color': grid_color,
    'legend.facecolor': card_bg,
    'legend.edgecolor': border_color,
    'font.family': 'DejaVu Sans',
    'axes.titlesize': 13,
    'axes.labelsize': 12,
    'legend.fontsize': 11,
})


def style_axes(ax, title, x_max):
    """Style axes with singularity marker"""
    ax.set_title(title, fontweight="bold", color=accent_color)
    ax.set_xlabel("x", color=text_color)
    ax.set_ylabel("y", color=text_color)
    ax.grid(True, alpha=0.25, color=grid_color)
    ax.set_xlim(0.0, x_max)
    
    # Red singularity marker at x = π/2
    ax.axvline(
        x=math.pi/2,
        color=COLOR_PALETTES["sing"],
        linestyle="--",
        linewidth=2.0,
        alpha=0.85,
        label="singularity π/2"
    )
    ax.tick_params(colors=text_color)


def autoscale_y_with_padding(ax, pad_frac=0.10):
    """Auto-scale y-axis with padding"""
    ys = []
    for line in ax.get_lines():
        y = np.asarray(line.get_ydata(), dtype=float)
        y = y[np.isfinite(y)]
        if y.size:
            ys.append(y)
    if not ys:
        return
    y_all = np.concatenate(ys)
    ymin, ymax = float(np.min(y_all)), float(np.max(y_all))
    if math.isclose(ymin, ymax, rel_tol=1e-12, abs_tol=1e-12):
        ymin -= 1.0
        ymax += 1.0
    span = ymax - ymin
    ax.set_ylim(ymin - pad_frac * span, ymax + pad_frac * span)


def plot_method_once(ax, a, method_key, h, tol, label, left_frac, right_frac, x_domain_max, color_idx=0):
    """Plot a single method solution with specified color"""
    x, y = solve_combined(
        a, method_key, h=h, tol=tol,
        left_frac=left_frac, right_frac=right_frac, 
        x_domain_max=x_domain_max, gap_nan=True
    )
    color = COLOR_PALETTES[method_key][color_idx % len(COLOR_PALETTES[method_key])]
    ax.plot(x, y, lw=2.6, color=color, label=label)
    return x, y


def plot_exact_once(ax, a, left_frac, right_frac, x_domain_max):
    """Plot exact solution"""
    x, y = exact_combined(
        a, left_frac=left_frac, right_frac=right_frac, 
        x_domain_max=x_domain_max, n=1400
    )
    ax.plot(x, y, lw=2.2, linestyle="--", color=COLOR_PALETTES["exact"], label="Exact")
    return x, y


def relative_error_stats(x, y, a, eps=1e-12):
    """Compute relative error statistics"""
    m = np.isfinite(x) & np.isfinite(y)
    if not np.any(m):
        return float("nan"), float("nan"), float("nan")
    xf = x[m]
    yf = y[m]
    ye = y_exact(xf, a)
    r = np.abs(yf - ye) / (np.abs(ye) + eps)
    return float(np.max(r)), float(np.mean(r)), float(np.sqrt(np.mean(r**2)))


def export_excel_derivative_tables(a: float, methods_to_export: list, h=0.05, tol=1e-6, 
                                   left_frac=0.99, right_frac=1.01, x_domain_max=math.pi):
    """Export derivative tables to Excel"""
    if not EXCEL_AVAILABLE:
        return None, "Excel export requires openpyxl. Install with: pip install openpyxl"
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Domain grids
    x_left_end = left_frac * math.pi / 2.0
    x_right_start = right_frac * math.pi / 2.0
    nL, nR = 600, 600
    xL_grid = np.linspace(0.0, x_left_end, nL)
    xR_grid = np.linspace(x_right_start, x_domain_max, nR)
    
    # Exact derivatives
    ypL_ex = yprime_exact(xL_grid, a)
    ypR_ex = yprime_exact(xR_grid, a)
    
    pretty = {"euler": "Euler", "rk4": "RK4", "rk45": "RK45"}
    
    for method_key in methods_to_export:
        sheet_name = f"{pretty.get(method_key, method_key)}_Deriv"
        ws = wb.create_sheet(sheet_name)
        
        ws.append(["x", "yprime_exact", f"yprime_{method_key}", "error", "relative_error"])
        
        # Style header
        fill = PatternFill("solid", fgColor="EFE6D8")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Solve and compute derivatives
        xL, yL = solve_branch(a, method_key, branch="left", h=h, tol=tol,
                             left_frac=left_frac, right_frac=right_frac, 
                             x_domain_max=x_domain_max)
        xR, yR = solve_branch(a, method_key, branch="right", h=h, tol=tol,
                             left_frac=left_frac, right_frac=right_frac, 
                             x_domain_max=x_domain_max)
        
        yL_i = np.interp(xL_grid, xL, yL)
        yR_i = np.interp(xR_grid, xR, yR)
        
        ypL_m = yprime_from_y(yL_i, a)
        ypR_m = yprime_from_y(yR_i, a)
        
        errL = np.abs(ypL_m - ypL_ex)
        errR = np.abs(ypR_m - ypR_ex)
        relL = errL / (np.abs(ypL_ex) + 1e-12)
        relR = errR / (np.abs(ypR_ex) + 1e-12)
        
        # Write data
        for i in range(nL):
            ws.append([float(xL_grid[i]), float(ypL_ex[i]), float(ypL_m[i]), 
                      float(errL[i]), float(relL[i])])
        ws.append([None, None, None, None, None])  # gap
        for i in range(nR):
            ws.append([float(xR_grid[i]), float(ypR_ex[i]), float(ypR_m[i]), 
                      float(errR[i]), float(relR[i])])
        
        # Format numbers
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=5, max_row=ws.max_row):
            for c in row:
                if c.value is None:
                    continue
                if c.column == 1:
                    c.number_format = "0.000000"
                else:
                    c.number_format = "0.000000E+00"
        
        ws.freeze_panes = "A2"
    
    # README sheet
    wsR = wb.create_sheet("README")
    wsR.append(["Export type: Derivative tables"])
    wsR.append([f"a = {a}"])
    wsR.append([f"methods = {methods_to_export}"])
    wsR.append(["Columns per method: x, yprime_exact, yprime_method, error, relative_error"])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, None


def main():
    """Main Streamlit application"""
    
    # Sidebar controls
    with st.sidebar:
        st.header("⚙️ Parameters")
        
        # Domain controls
        st.subheader("Domain Settings")
        left_frac = st.slider("Left Fraction (default 0.99)", 
                             min_value=0.5, max_value=0.999, value=0.99, step=0.001,
                             help="Fraction of π/2 for left domain boundary")
        right_frac = st.slider("Right Fraction (default 1.01)", 
                              min_value=1.001, max_value=1.5, value=1.01, step=0.001,
                              help="Fraction of π/2 for right domain boundary")
        x_domain_max = st.slider("Domain Max (π)", 
                                min_value=1.0, max_value=4.0, value=math.pi, step=0.1,
                                format="%.3f",
                                help="Maximum x value in domain")
        
        st.divider()
        
        # Numerical method controls
        st.subheader("Numerical Settings")
        h_size = st.number_input("Step Size (h)", 
                                min_value=0.001, max_value=0.1, value=0.05, step=0.001,
                                format="%.3f",
                                help="Step size for Euler and RK4 methods")
        tol = st.number_input("Tolerance (RK45)", 
                             min_value=1e-10, max_value=1e-3, value=1e-6, step=1e-7,
                             format="%.0e",
                             help="Tolerance for RK45 adaptive method")
    
    # Compact header with LaTeX
    col_header1, col_header2, col_header3 = st.columns([1, 2, 1])
    with col_header2:
        st.markdown('<p class="main-header">⚡ RICCATI BVP EXPLORER</p>', unsafe_allow_html=True)
        st.markdown('<p class="sub-header">Modern Numerical Methods Suite</p>', unsafe_allow_html=True)
        
        # LaTeX equation display
        st.markdown(f'<div class="equation-box">', unsafe_allow_html=True)
        
        st.latex(r"y'(x) = a \cdot y(x)^2 + \frac{1}{a}")
        st.caption("Exact solution: " + r"$y(x) = \frac{1}{a} \tan(x)$" + f" | Singularity at " + r"$x = \frac{\pi}{2}$")
        left_str = f"{left_frac:.2f}"
        right_str = f"{right_frac:.2f}"
        domain_str = f"{x_domain_max:.2f}"
        st.caption(f"Domain: " + r"$[0, " + left_str + r"\cdot\frac{\pi}{2}] \cup [" + right_str + r"\cdot\frac{\pi}{2}, " + domain_str + r"]$")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Calculate domain boundaries
    x_left_end = left_frac * math.pi / 2
    x_right_start = right_frac * math.pi / 2
    
    # Create tabs
    tab1, tab2, tab3 = st.tabs(["🧠 Solution", "📊 Comparing Methods", "🎚️ Slider (a)"])
    
    # ================= Tab 1: Solution =================
    with tab1:
        st.subheader("Single Method Solution")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            a1 = st.number_input("Parameter a", value=2.0, min_value=0.1, max_value=10.0, 
                                step=0.1, key="a1", format="%.2f")
        
        with col2:
            method1 = st.selectbox("Method", ["Euler", "RK4", "RK45"], key="method1")
        
        with col3:
            show_exact1 = st.checkbox("Show Exact", value=True, key="exact1")
        
        with col4:
            use_x1 = st.checkbox("Use x marker", value=True, key="usex1")
        
        x1_val = None
        if use_x1:
            x1_val = st.number_input("x value", value=1.0, min_value=0.0, 
                                    max_value=float(x_domain_max), step=0.1, key="x1")
        
        autofocus1 = st.checkbox("Focus near x", value=True, key="focus1")
        
        # Plot button
        if st.button("⚡ Plot Solution", type="primary", use_container_width=True):
            if abs(a1) < 1e-12:
                st.error("Parameter a must be nonzero.")
            else:
                method_map = {"Euler": "euler", "RK4": "rk4", "RK45": "rk45"}
                method_key = method_map[method1]
                
                fig, ax = plt.subplots(figsize=(14, 7))
                
                x, y = plot_method_once(ax, a1, method_key, h_size, tol, method1, 
                                       left_frac, right_frac, x_domain_max, color_idx=0)
                if show_exact1:
                    plot_exact_once(ax, a1, left_frac, right_frac, x_domain_max)
                
                # Add x marker if specified
                if use_x1 and x1_val is not None:
                    if 0 <= x1_val <= x_domain_max and not (x_left_end < x1_val < x_right_start):
                        y_at_x = float(np.interp(x1_val, x, y))
                        ax.plot(x1_val, y_at_x, 'rv', markersize=10, markeredgewidth=2,
                               label=f'x={x1_val:.3f}', zorder=10)
                
                style_axes(ax, f"{method1} Solution (a={a1})", x_domain_max)
                ax.legend(loc="best", framealpha=0.9)
                autoscale_y_with_padding(ax, pad_frac=0.10)
                
                if use_x1 and x1_val is not None and autofocus1:
                    lo = max(0.0, x1_val - 0.45)
                    hi = min(x_domain_max, x1_val + 0.45)
                    ax.set_xlim(lo, hi)
                
                plt.tight_layout()
                st.pyplot(fig)
                plt.close(fig)
                
                # Status info
                info_lines = [
                    f"a={a1} | method={method1}",
                    f"domain=[0, {left_frac:.3f}·π/2] ∪ [{right_frac:.3f}·π/2, {x_domain_max:.3f}]",
                    f"h={h_size:.3f} | tol={tol:.0e}"
                ]
                
                if use_x1 and x1_val is not None:
                    if x_left_end < x1_val < x_right_start:
                        info_lines.append("x is inside forbidden gap near π/2 → skipped")
                    elif x1_val < 0 or x1_val > x_domain_max:
                        info_lines.append(f"x outside [0, {x_domain_max:.3f}] → skipped")
                    else:
                        branch = "left" if x1_val <= x_left_end else "right"
                        xb, yb = solve_branch(
                            a1, method_key, branch=branch, h=h_size, tol=tol,
                            left_frac=left_frac, right_frac=right_frac, 
                            x_domain_max=x_domain_max
                        )
                        y_num = float(np.interp(x1_val, xb, yb))
                        info_lines.append(f"x={x1_val:.6f} | y_num={y_num:.6e}")
                        if show_exact1:
                            y_ex = float(y_exact(np.array([x1_val]), a1)[0])
                            err_abs = abs(y_num - y_ex)
                            err_rel = err_abs / (abs(y_ex) + 1e-12)
                            info_lines.append(f"y_exact={y_ex:.6e} | |err|={err_abs:.3e} | rel={err_rel:.3e}")
                
                st.markdown(f'<div class="info-box">{"<br>".join(info_lines)}</div>', 
                           unsafe_allow_html=True)
        
        # Excel export
        if st.button("📑 Export Excel (Derivative Table)", use_container_width=True, key="export1"):
            if abs(a1) < 1e-12:
                st.error("a must be nonzero.")
            else:
                method_map = {"Euler": "euler", "RK4": "rk4", "RK45": "rk45"}
                methods_to_export = [method_map[method1]]
                
                output, error = export_excel_derivative_tables(a1, methods_to_export, h=h_size, 
                                                              tol=tol, left_frac=left_frac, 
                                                              right_frac=right_frac, 
                                                              x_domain_max=x_domain_max)
                if error:
                    st.error(error)
                else:
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"riccati_deriv_a{a1:.1f}_{ts}.xlsx"
                    st.download_button(
                        label="📥 Download Excel File",
                        data=output,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl1"
                    )
    
    # ================= Tab 2: Comparing Methods =================
    with tab2:
        st.subheader("Method Comparison")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            a2 = st.number_input("Parameter a", value=2.0, min_value=0.1, max_value=10.0, 
                                step=0.1, key="a2", format="%.2f")
        
        with col2:
            x2_val = st.text_input("x (optional)", value="", key="x2", 
                                   help="Optional x value for marker")
        
        with col3:
            show_exact2 = st.checkbox("Show Exact", value=True, key="exact2")
        
        # Auto-plot on parameter change
        if abs(a2) >= 1e-12:
            x_val = None
            if x2_val.strip():
                try:
                    x_val = float(x2_val)
                    if x_val < 0 or x_val > x_domain_max:
                        st.error(f"x must be in [0, {x_domain_max:.3f}]")
                        x_val = None
                except:
                    x_val = None
            
            # Solution plots (side by side) - BIG graphs
            fig_solutions, axes = plt.subplots(1, 3, figsize=(20, 6))
            
            methods_data = [
                ("euler", "Euler", axes[0], 0),
                ("rk4", "RK4", axes[1], 0),
                ("rk45", "RK45", axes[2], 0)
            ]
            
            results = {}
            
            for method_key, method_name, ax, color_idx in methods_data:
                x, y = plot_method_once(ax, a2, method_key, h_size, tol, method_name, 
                                       left_frac, right_frac, x_domain_max, color_idx=color_idx)
                if show_exact2:
                    plot_exact_once(ax, a2, left_frac, right_frac, x_domain_max)
                
                if x_val is not None:
                    if 0 <= x_val <= x_domain_max and not (x_left_end < x_val < x_right_start):
                        y_at_x = float(np.interp(x_val, x, y))
                        ax.plot(x_val, y_at_x, 'rv', markersize=10, markeredgewidth=2,
                               label=f'x={x_val:.3f}', zorder=10)
                
                style_axes(ax, f"{method_name} Solution", x_domain_max)
                ax.legend(loc="best", framealpha=0.9)
                autoscale_y_with_padding(ax, pad_frac=0.10)
                
                results[method_key] = (x, y)
            
            plt.tight_layout()
            st.pyplot(fig_solutions)
            plt.close(fig_solutions)
            
            # Combined error plot - BIG graph
            fig_err, ax_err = plt.subplots(figsize=(14, 7))
            
            maxr_e = 0.0
            maxr_4 = 0.0
            maxr_5 = 0.0
            
            color_idx = 0
            for method_key, method_name in [("euler", "Euler"), ("rk4", "RK4"), ("rk45", "RK45")]:
                x, y = results[method_key]
                m = np.isfinite(x) & np.isfinite(y)
                if np.any(m):
                    xf = x[m]
                    yf = y[m]
                    ye = y_exact(xf, a2)
                    r = np.abs(yf - ye) / (np.abs(ye) + 1e-12)
                    color = COLOR_PALETTES[method_key][color_idx % len(COLOR_PALETTES[method_key])]
                    ax_err.plot(xf, r, lw=2.5, color=color, label=method_name)
                    max_r = float(np.max(r))
                    if method_key == "euler":
                        maxr_e = max_r
                    elif method_key == "rk4":
                        maxr_4 = max_r
                    else:
                        maxr_5 = max_r
                color_idx += 1
            
            ax_err.set_yscale('log')
            style_axes(ax_err, "Relative Errors Comparison", x_domain_max)
            ax_err.set_ylabel("Relative Error r(x)")
            ax_err.legend(loc="best", framealpha=0.9)
            plt.tight_layout()
            st.pyplot(fig_err)
            plt.close(fig_err)
            
            # Info box
            info_lines = [
                f"a = {a2}",
                f"Domain: [0, {left_frac:.3f}·π/2] ∪ [{right_frac:.3f}·π/2, {x_domain_max:.3f}]",
                f"h={h_size:.3f} | tol={tol:.0e}",
                ""
            ]
            
            if x_val is not None:
                info_lines.extend([
                    f"At x = {x_val:.6f}:",
                    "Method    | y_num      | y_exact    | Error      | Rel Error",
                    "----------|------------|------------|------------|----------"
                ])
                
                y_exact_val = float(y_exact(np.array([x_val]), a2)[0])
                
                for method_key, method_name in [("euler", "Euler"), ("rk4", "RK4"), ("rk45", "RK45")]:
                    try:
                        x, y = solve_combined(a2, method_key, h=h_size, tol=tol,
                                             left_frac=left_frac, right_frac=right_frac,
                                             x_domain_max=x_domain_max, gap_nan=False)
                        y_num_val = float(np.interp(x_val, x, y))
                        error = abs(y_num_val - y_exact_val)
                        rel_error = error / (abs(y_exact_val) + 1e-12)
                        info_lines.append(
                            f"{method_name:<9} | {y_num_val:>10.6e} | {y_exact_val:>10.6e} | {error:>10.3e} | {rel_error:>8.3e}"
                        )
                    except Exception as e:
                        info_lines.append(f"{method_name:<9} | Error: {str(e)}")
                
                info_lines.append("")
            
            info_lines.extend([
                "Global Relative Error Statistics:",
                "Method    | Max Error",
                "----------|----------",
                f"Euler     | {maxr_e:>8.3e}",
                f"RK4       | {maxr_4:>8.3e}",
                f"RK45      | {maxr_5:>8.3e}"
            ])
            
            st.markdown(f'<div class="info-box">{"<br>".join(info_lines)}</div>', 
                       unsafe_allow_html=True)
        
        # Excel export for Tab 2
        if st.button("📑 Export Excel (All Methods)", use_container_width=True, key="export2"):
            if abs(a2) < 1e-12:
                st.error("a must be nonzero.")
            else:
                methods_to_export = ["euler", "rk4", "rk45"]
                output, error = export_excel_derivative_tables(a2, methods_to_export, h=h_size, 
                                                              tol=tol, left_frac=left_frac, 
                                                              right_frac=right_frac, 
                                                              x_domain_max=x_domain_max)
                if error:
                    st.error(error)
                else:
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"riccati_deriv_all_methods_a{a2:.1f}_{ts}.xlsx"
                    st.download_button(
                        label="📥 Download Excel File",
                        data=output,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl2"
                    )
    
    # ================= Tab 3: Slider (a) =================
    with tab3:
        st.subheader("Interactive Parameter Slider")
        
        col1, col2, col3 = st.columns([2, 1.5, 2])
        
        with col1:
            multi_a_mode = st.checkbox("Multi-a Mode", value=False, 
                                      help="Plot multiple a values on same axes")
        
        with col2:
            if not multi_a_mode:
                a3 = st.slider("Parameter a", min_value=1, max_value=10, value=2, 
                              step=1, key="a3_slider")
                a3_val = float(a3)
                st.write(f"**Current a: {a3_val:.1f}**")
            else:
                a_multi_input = st.text_input("a values (comma-separated)", 
                                             value="-5,-2,-1,1,2,5", key="a_multi")
                try:
                    a_values = [float(s.strip()) for s in a_multi_input.split(',') if s.strip()]
                    a3_val = a_values[0] if a_values else 2.0
                except:
                    st.error("Invalid a values format. Use comma-separated numbers.")
                    a_values = [2.0]
                    a3_val = 2.0
        
        with col3:
            method3 = st.selectbox("Method", ["Euler", "RK4", "RK45", "All"], key="method3")
            show_exact3 = st.checkbox("Show Exact", value=True, key="exact3")
        
        # BIG graph - Auto-update on parameter change
        fig3, ax3 = plt.subplots(figsize=(14, 7))
        
        if multi_a_mode:
            colors = ['#2563EB', '#10B981', '#F59E0B', '#DC2626', '#8B5CF6', '#EC4899', '#06B6D4']
            
            if method3 == "Euler":
                methods = [("euler", "Euler")]
            elif method3 == "RK4":
                methods = [("rk4", "RK4")]
            elif method3 == "RK45":
                methods = [("rk45", "RK45")]
            else:
                methods = [("euler", "Euler"), ("rk4", "RK4"), ("rk45", "RK45")]
            
            for i, a_val in enumerate(a_values):
                if abs(a_val) < 1e-12:
                    continue
                color = colors[i % len(colors)]
                
                method_color_idx = 0
                for key, name in methods:
                    label = f"{name} (a={a_val})" if len(methods) > 1 else f"a={a_val}"
                    x, y = solve_combined(
                        a_val, key, h=h_size, tol=tol,
                        left_frac=left_frac, right_frac=right_frac, 
                        x_domain_max=x_domain_max, gap_nan=True
                    )
                    # Use different color for each method when "All" is selected
                    if len(methods) > 1:
                        plot_color = COLOR_PALETTES[key][method_color_idx % len(COLOR_PALETTES[key])]
                        ax3.plot(x, y, lw=2.6, color=plot_color, linestyle='-', label=label)
                        method_color_idx += 1
                    else:
                        ax3.plot(x, y, lw=2.6, color=color, linestyle='-', label=label)
                
                if show_exact3:
                    x, y = exact_combined(
                        a_val, left_frac=left_frac, right_frac=right_frac, 
                        x_domain_max=x_domain_max, n=1400
                    )
                    ax3.plot(x, y, lw=2.2, linestyle=":", color=color, alpha=0.7,
                            label=f"Exact (a={a_val})" if len(a_values) > 1 else "Exact")
        else:
            if method3 == "Euler":
                methods = [("euler", "Euler")]
            elif method3 == "RK4":
                methods = [("rk4", "RK4")]
            elif method3 == "RK45":
                methods = [("rk45", "RK45")]
            else:
                methods = [("euler", "Euler"), ("rk4", "RK4"), ("rk45", "RK45")]
            
            color_idx = 0
            for key, name in methods:
                plot_method_once(ax3, a3_val, key, h_size, tol, name, 
                               left_frac, right_frac, x_domain_max, color_idx=color_idx)
                color_idx += 1
            
            if show_exact3:
                plot_exact_once(ax3, a3_val, left_frac, right_frac, x_domain_max)
        
        style_axes(ax3, f"{method3} — Fixed Scale", x_domain_max)
        ax3.legend(loc="best", framealpha=0.9)
        autoscale_y_with_padding(ax3, pad_frac=0.10)
        plt.tight_layout()
        st.pyplot(fig3)
        plt.close(fig3)
        
        # Info box for Tab 3
        if multi_a_mode:
            info_lines = ["Relative Error Statistics for Different a Values:", 
                         "a Value   | Max Error", "----------|----------"]
            for a_val in a_values:
                if abs(a_val) < 1e-12:
                    continue
                try:
                    x, y = solve_combined(a_val, "rk45", h=h_size, tol=tol,
                                         left_frac=left_frac, right_frac=right_frac,
                                         x_domain_max=x_domain_max, gap_nan=False)
                    maxr, _, _ = relative_error_stats(x, y, a_val, eps=1e-12)
                    info_lines.append(f"{a_val:>8.1f} | {maxr:>8.3e}")
                except:
                    info_lines.append(f"{a_val:>8.1f} | Error")
            st.markdown(f'<div class="info-box">{"<br>".join(info_lines)}</div>', 
                       unsafe_allow_html=True)
        
        # Excel export for Tab 3
        if st.button("📑 Export Excel", use_container_width=True, key="export3"):
            if multi_a_mode:
                if not a_values:
                    st.error("No valid a values provided")
                else:
                    if method3 == "Euler":
                        methods_to_export = ["euler"]
                    elif method3 == "RK4":
                        methods_to_export = ["rk4"]
                    elif method3 == "RK45":
                        methods_to_export = ["rk45"]
                    else:
                        methods_to_export = ["euler", "rk4", "rk45"]
                    
                    # Export one file per a value
                    for a_val in a_values:
                        if abs(a_val) < 1e-12:
                            continue
                        output, error = export_excel_derivative_tables(a_val, methods_to_export, 
                                                                      h=h_size, tol=tol, 
                                                                      left_frac=left_frac, 
                                                                      right_frac=right_frac, 
                                                                      x_domain_max=x_domain_max)
                        if error:
                            st.error(f"Error for a={a_val}: {error}")
                        else:
                            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                            filename = f"riccati_deriv_a{a_val:.1f}_{ts}.xlsx"
                            st.download_button(
                                label=f"📥 Download Excel (a={a_val:.1f})",
                                data=output,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl3_{a_val}"
                            )
            else:
                if abs(a3_val) < 1e-12:
                    st.error("a must be nonzero.")
                else:
                    if method3 == "Euler":
                        methods_to_export = ["euler"]
                    elif method3 == "RK4":
                        methods_to_export = ["rk4"]
                    elif method3 == "RK45":
                        methods_to_export = ["rk45"]
                    else:
                        methods_to_export = ["euler", "rk4", "rk45"]
                    
                    output, error = export_excel_derivative_tables(a3_val, methods_to_export, 
                                                                  h=h_size, tol=tol, 
                                                                  left_frac=left_frac, 
                                                                  right_frac=right_frac, 
                                                                  x_domain_max=x_domain_max)
                    if error:
                        st.error(error)
                    else:
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"riccati_deriv_a{a3_val:.1f}_{ts}.xlsx"
                        st.download_button(
                            label="📥 Download Excel File",
                            data=output,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl3"
                        )


if __name__ == "__main__":
    main()
